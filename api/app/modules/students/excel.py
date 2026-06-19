"""Excel parsing helper (openpyxl)."""
from __future__ import annotations

import io
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def parse_preview(file_bytes: bytes) -> dict:
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return {"columns": [], "row_count": 0}
    headers = [str(c) if c is not None else None for c in header_row]
    sample_rows = []
    for r in rows_iter:
        sample_rows.append([str(c) if c is not None else None for c in r])
        if len(sample_rows) >= 5:
            break
    columns = []
    for i, h in enumerate(headers):
        letter = get_column_letter(i + 1)
        samples = []
        for r in sample_rows:
            if i < len(r):
                samples.append(r[i])
        columns.append({"letter": letter, "header": h, "sample_values": samples})
    # Total rows is approximate (openpyxl read_only doesn't know without scanning)
    # We re-open and count
    wb2 = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws2 = wb2.active
    total = 0
    for _ in ws2.iter_rows(min_row=2, values_only=True):
        total += 1
    return {"columns": columns, "row_count": total}


def parse_rows(file_bytes: bytes) -> list[list[Any]]:
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return []
    headers = [str(c) if c is not None else None for c in header_row]
    out = []
    for r in rows_iter:
        row = list(r)
        # Normalize None → ""
        norm = [str(c) if c is not None else "" for c in row]
        out.append(norm)
    return [headers] + out
